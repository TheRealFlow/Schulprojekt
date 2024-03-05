import csv
from typing import List
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from os import path
from faker import Faker
from random import choice
from versandplanung.articles.helpers import get_random_selection_from_list

from versandplanung.articles.models import Article
from versandplanung.customer.models import Customer
from versandplanung.orders.models import Orders


class Command(BaseCommand):
    help = "Generates new orders overriding the existing ones."

    def handle(self, *args, **kwargs):
        fake = Faker()
        customer_data: List[Customer] = []
        try:
            file_path = path.abspath('datasources/customer_data.csv')
            with open(file_path, 'r') as csv_file:
                csv_reader = csv.DictReader(csv_file)
                for row in csv_reader:
                    customer_data.append(Customer(
                        row["UserId"], row["First_Name"], row["Last_Name"], row["City"], row['Street'], row['Street_Number'], row["E-Mail"]))
            csv_file.close()
        except FileNotFoundError as error:
            raise CommandError("Couldn't find customer_data.csv file.", error)

        article_data: List[Article] = []
        try:
            articles_workbook = load_workbook('datasources/articles.xlsx')
            article_sheet = articles_workbook.active

            for row in article_sheet.iter_rows(min_row=1, max_row=article_sheet.max_row, values_only=True):
                article_data.append(
                    Article(id=row[0], price=row[1], name=row[2], short_description=row[3]))

            articles_workbook.close()
        except FileNotFoundError:
            raise CommandError("Couldn't find articles.xlsx file.")

        Orders.objects.all().delete()
        for user in customer_data:
            random_articles = get_random_selection_from_list(article_data)
            # black magic: don't touch
            articles = list(filter(lambda x: articleFilter(
                list(map(lambda x: list(x.keys())[0], random_articles)), x), article_data))
            article_prices = {}
            for item in articles:
                article_prices[item.get_name()] = item.get_price()

            total_sum = 0
            for article_sum in random_articles:
                try:
                    price_per_unit = int(
                        article_prices[list(article_sum.keys())[0]])
                except ValueError:
                    price_per_unit = 0
                total_sum += price_per_unit * \
                    int(list(article_sum.values())[0])

            Orders.objects.create(
                orderNumber=fake.uuid4(),
                customerId=user.get_id(),
                articles=random_articles,
                address=user.get_city() + ", " + user.get_street() +
                " " + str(user.get_street_number()),
                status=choice(Orders.STATUS_CHOICES)[0],
                total=total_sum
            )


def articleFilter(search_vals: List[str], val: Article) -> bool:
    if val.get_name() in search_vals:
        return True
    return False

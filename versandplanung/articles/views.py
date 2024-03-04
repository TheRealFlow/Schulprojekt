from os import path
from versandplanung.settings import BASE_DIR
from django.shortcuts import render
from openpyxl import load_workbook

from versandplanung.articles.models import Article


def view_articles(request):
    articles = []
    
    # we need the base dir in the path to make it run in the exe since the path is random their
    article_path = "datasources/articles.xlsx"
    absolute_path = path.join(BASE_DIR, article_path)
    articles_workbook = load_workbook(absolute_path)
    article_sheet = articles_workbook.active

    # min_row=2 because the first row contains the headers
    for row in article_sheet.iter_rows(min_row=2, max_row=article_sheet.max_row, values_only=True):
        articles.append(
            Article(id=row[0], price=row[1], name=row[2], short_description=row[3]))

    articles_workbook.close()

    return render(request, 'articles.html', {
        'articles': articles
    })
